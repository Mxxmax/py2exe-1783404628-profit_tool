#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
盈利诊断报表 通用数据处理 — Excel公式计算列
=============================================
读取当前目录的 盈利诊断报表_数据.xlsx，输出 盈利诊断报表_计算结果.xlsx
自动检测月份、"-"替换为0，计算列浅蓝底色，列顺序按指定模板。
"""

import re
import os
import sys
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(sys.executable)) if getattr(sys, 'frozen', False) else os.getcwd()
INPUT_FILE = os.path.join(SCRIPT_DIR, '盈利诊断报表_数据.xlsx')
if not os.path.exists(INPUT_FILE):
    print(f'错误: 找不到 {INPUT_FILE}')
    print('请将文件 盈利诊断报表_数据.xlsx 放在当前目录')
    sys.exit(1)
OUTPUT_FILE = os.path.join(SCRIPT_DIR, '盈利诊断报表_计算结果.xlsx')

# ========== 1. 读取数据 ==========
print(f'读取: {INPUT_FILE}')
wb = load_workbook(INPUT_FILE, data_only=True)
ws = wb.active
row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]

# 自动检测月份
available_months = set()
for j in range(5, len(row5)):
    v = row5[j]
    if v is not None and '月' in str(v) and '当年' not in str(v):
        available_months.add(str(v).strip())
available_months = sorted(available_months, key=lambda x: int(x.replace('月','')))
MONTH = available_months[0]
print(f'可用月份: {available_months} → 使用: {MONTH}')

# 收集三级指标
sanji_rows = []
for i, vals in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=4, values_only=True), 6):
    a, b, c, d = vals
    if c is not None and b is not None:
        sanji_rows.append((i, b, c, d))
sanji_names = [r[2] for r in sanji_rows]
print(f'三级指标: {len(sanji_names)}')

# 找到目标月份的公司列
month_cols = []
j = 5
while j < len(row4):
    cname = row4[j]
    if cname is not None:
        label_may = row5[j] if j < len(row5) else None
        if label_may == MONTH:
            month_cols.append((j, j+1, cname))
        j += 2
    else:
        j += 1
print(f'{MONTH} 公司数: {len(month_cols)}')

# 构建宽表
records = []
for col_may, col_acc, company_name in month_cols:
    record = {'经销商名称': company_name}
    for row_num, erji_name, sanji_name, jixing in sanji_rows:
        val = ws.cell(row=row_num, column=col_may + 1).value
        record[sanji_name] = val
    records.append(record)
print(f'有效记录: {len(records)}')

# ========== 2. 公式定义 ==========
# 公式内部名 → (输出显示名, 公式表达式)
FORMULA_DEFS_INTERNAL = [
    ('总体收入计算', '总体收入', '总体(新车收入)+总体(服务产值,以下均不含续保)+二手车&置换收入+总体（增值收入）+其他业务收入'),
    ('新车毛利计算', '新车毛利', '总体(新车收入)-总体新车成本（新车收入*96%）'),
    ('平均单价计算', '平均单价', '总体(新车收入)/新车销售台次'),
    ('平均毛利计算', '平均毛利', '新车毛利/新车销售台次'),
    ('二手车&置换收入计算', '二手车&置换收入', '二手车销售台次*二手车销售平均单价+二手车中介台次*二手车中介平均收入'),
    ('二手车&置换毛利计算', '二手车&置换毛利', '二手车&置换收入-总体二手车成本（二手车收入*85%）'),
    ('二手车单车平均毛利计算', '二手车单车平均毛利', '二手车&置换毛利/(二手车销售台次+二手车中介台次)'),
    ('增值毛利计算', '增值毛利', '总体（增值收入）-总体增值业务成本（增值收入*13%）'),
    ('总体返利计算', '总体返利', '销售运营与满意度返利+增值返利（如金融）+售后运营返利+备件精品返利+二手车运营返利+一网建店返利'),
    ('总体成本计算', '总体成本', '总体新车成本（新车收入*96%）+总体售后服务成本（服务收入*55%）+总体二手车成本（二手车收入*85%）+总体增值业务成本（增值收入*13%）+其他业务成本'),
    ('总体费用计算', '总体费用', '店面费用+财务费用+人力费用+广宣费用+运营费用'),
    ('总体利润计算', '总体利润', '总体收入计算+总体返利计算-总体成本计算-总体费用计算'),
    ('人均薪酬计算', '人均薪酬', '人力费用/总人数'),
    # ===== 新增：原数据列→公式替换（工作簿黄底行）=====
    ('总体（增值收入）计算', '总体（增值收入）',
     '(个贷-金融服务台次（不含二网）-个贷-非贴息金融服务台次（不含二网）)*个贷-贴息单车金融服务费+个贷-非贴息金融服务台次（不含二网）*个贷-非贴息单车金融服务费（含返佣）+保险-新车保险台次（不含二网）*保险-新车保险单车返点+牌证-上牌服务费总收入+精品-新车加装精品台次*精品-新车单车精品收入+自营非车险-销售单量*自营非车险-单车销售收入+三方非车险-销售单量*三方非车险-单车返佣+太阳膜-销售台次*太阳膜-单车销售收入'),
    ('总体新车成本（新车收入*96%）计算', '总体新车成本（新车收入*96%）', '新车促销成本+新车采购成本'),
    ('总体二手车成本（二手车收入*85%）计算', '总体二手车成本（二手车收入*85%）', '二手车采购成本+二手车整备成本'),
    ('总体增值业务成本（增值收入*13%）计算', '总体增值业务成本（增值收入*13%）', '代办上牌成本+精品加装成本+自营非车险-总成本+太阳膜-总成本'),
    ('店面费用计算', '店面费用', '店面租金+店面折旧费用+店面摊销费用'),
    ('财务费用计算', '财务费用', '融资费用+手续费用'),
    ('广宣费用计算', '广宣费用', '网销平台费用+新媒体费用+企划活动费用'),
    ('运营费用计算', '运营费用', '差旅费+通讯费+招待费+办公费+水电费+汽车费用+其他'),
]

# 构建查找表：输出名→内部名, 输出名→公式表达式
FORMULA_OUTPUT_NAMES = {f[1] for f in FORMULA_DEFS_INTERNAL}
formula_output_to_internal = {f[1]: f[0] for f in FORMULA_DEFS_INTERNAL}
formula_output_to_expr = {f[1]: f[2] for f in FORMULA_DEFS_INTERNAL}
formula_internal_to_expr = {f[0]: f[2] for f in FORMULA_DEFS_INTERNAL}

# ========== 3. 定义输出列顺序 ==========
OUTPUT_COLUMNS = [
    '经销商名称',
    '总体利润', '综合毛利率', '利润率',
    '总体收入', '总体(新车收入)', '新车毛利', '新车销售台次', '平均单价', '平均毛利',
    '展厅销量占比', '网销销量占比', '新媒体销量占比', '二网销量占比（含批售）',
    '总体(服务产值,以下均不含续保)', '服务毛利', '维修台次', '客单价', '单车毛利',
    '工时收入占比', '机修收入占比', '事故车（钣喷）收入占比', '保养收入占比', '索赔收入占比', '备件（批发）收入占比',
    '二手车&置换收入', '二手车&置换毛利', '置换台次', '二手车销售台次', '二手车销售平均单价', '二手车单车平均毛利', '二手车中介台次', '二手车中介平均收入',
    '总体（增值收入）', '增值毛利',
    '个贷-金融服务台次（不含二网）', '个贷-非贴息金融服务台次（不含二网）', '个贷-贴息单车金融服务费', '个贷-非贴息单车金融服务费（含返佣）',
    '个贷-金融渗透率（不含二网）',
    '保险-新车保险台次（不含二网）', '保险-新车保险单车返点', '保险-新车保险渗透率',
    '牌证-上牌服务费总收入',
    '精品-新车加装精品台次', '精品-新车单车精品收入', '精品-精品加装渗透率（不含二网）',
    '自营非车险-销售单量', '自营非车险-单车销售收入',
    '三方非车险-销售单量', '三方非车险-单车返佣',
    '太阳膜-销售台次', '太阳膜-单车销售收入',
    '其他业务收入', '其他业务毛利',
    '总体返利', '销售运营与满意度返利', '增值返利（如金融）', '售后运营返利', '备件精品返利', '二手车运营返利', '一网建店返利', '形象焕新返利',
    '总体成本', '总体新车成本（新车收入*96%）', '新车采购成本', '新车促销成本', '总体售后服务成本（服务收入*55%）',
    '总体二手车成本（二手车收入*85%）', '二手车采购成本', '二手车整备成本',
    '总体增值业务成本（增值收入*13%）', '代办上牌成本', '精品加装成本', '自营非车险-总成本', '太阳膜-总成本', '其他业务成本',
    '总体费用', '店面费用', '店面租金', '店面折旧费用', '店面摊销费用',
    '财务费用', '融资费用', '手续费用',
    '人力费用', '人均薪酬', '总人数', '其中：销售人数', '售后人数',
    '广宣费用', '网销平台费用', '新媒体费用', '企划活动费用',
    '运营费用', '差旅费', '通讯费', '招待费', '办公费', '水电费', '汽车费用', '其他',
]

print(f'输出列数: {len(OUTPUT_COLUMNS)}')

# ========== 4. 构建列号映射 ==========
# 每个输出名对应的列字母
col_letter_of = {}
for ci, col_name in enumerate(OUTPUT_COLUMNS, 1):
    col_letter_of[col_name] = get_column_letter(ci)

# 公式内部名也映射到同一列
for internal_name, output_name, _ in FORMULA_DEFS_INTERNAL:
    if output_name in col_letter_of:
        col_letter_of[internal_name] = col_letter_of[output_name]

# 普通三级指标也映射
for mn in sanji_names:
    if mn not in col_letter_of:
        # 不在输出列清单中的指标不输出
        pass

# 判断每个输出列的类型
def is_formula_col(col_name):
    return col_name in FORMULA_OUTPUT_NAMES

# ========== 5. 预检公式引用 ==========
def check_formula_refs(formula_expr):
    all_names = set(sanji_names) | set(f[0] for f in FORMULA_DEFS_INTERNAL)
    sorted_all = sorted(all_names, key=len, reverse=True)
    remaining = formula_expr
    found_refs = []
    while remaining:
        matched = False
        for name in sorted_all:
            if remaining.startswith(name):
                found_refs.append(name)
                remaining = remaining[len(name):]
                matched = True
                break
        if not matched:
            if remaining[0] in '+-*/()（）\n\r\t ':
                remaining = remaining[1:]
                continue
            m = re.match(r'[\d.]+%?', remaining)
            if m:
                remaining = remaining[m.end():]
                continue
            return False, f"无法识别的片段: '{remaining[:20]}'"
    return True, found_refs

print('\n=== 公式引用验证 ===')
for internal_name, output_name, formula_expr in FORMULA_DEFS_INTERNAL:
    ok, result = check_formula_refs(formula_expr)
    if ok:
        print(f'  ✓ {output_name}')
    else:
        print(f'  ⚠ {output_name}: {result}')

# ========== 6. 写输出文件 ==========
header_font = Font(bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
calc_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

print(f'\n写入: {OUTPUT_FILE}')
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = f'{MONTH}计算结果'

# 收集所有需要用于公式替换的名称（原始指标名 + 公式内部名），按长度降序
all_replace_names = sorted(
    set(sanji_names) | set(f[0] for f in FORMULA_DEFS_INTERNAL),
    key=len, reverse=True
)

# 写表头
for ci, col_name in enumerate(OUTPUT_COLUMNS, 1):
    cell = out_ws.cell(row=1, column=ci, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    cell.border = thin_border

# 写数据
for ri, rec in enumerate(records, 2):
    for ci, col_name in enumerate(OUTPUT_COLUMNS, 1):
        cell = out_ws.cell(row=ri, column=ci)
        cell.border = thin_border
        
        if col_name == '经销商名称':
            cell.value = rec['经销商名称']
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            continue
        
        if is_formula_col(col_name):
            # 公式列：浅蓝底色
            cell.fill = calc_fill
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            formula_expr = formula_output_to_expr[col_name]
            expr = formula_expr
            for name in all_replace_names:
                if name in expr and name in col_letter_of:
                    expr = expr.replace(name, f'{col_letter_of[name]}{ri}')
            
            cell.value = f'={expr}'
            cell.number_format = '#,##0'
        else:
            # 普通数据列
            cell.alignment = Alignment(horizontal='right', vertical='center')
            raw_val = rec.get(col_name)
            
            if raw_val is None or raw_val == '':
                cell.value = 0
            elif isinstance(raw_val, str) and '%' in raw_val:
                cell.value = raw_val
            elif isinstance(raw_val, str) and raw_val in ('-', '—'):
                cell.value = 0
            elif isinstance(raw_val, str) and raw_val.replace(',','').replace(' ','').lstrip('-').replace('.','').isdigit():
                num_val = float(raw_val.replace(',','').replace(' ',''))
                cell.value = int(num_val) if num_val == int(num_val) else num_val
                cell.number_format = '#,##0'
            elif isinstance(raw_val, str) and '~' in raw_val:
                cell.value = 0
            else:
                cell.value = raw_val if raw_val else 0
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

# 列宽
out_ws.column_dimensions['A'].width = 42
for ci in range(2, len(OUTPUT_COLUMNS) + 1):
    out_ws.column_dimensions[get_column_letter(ci)].width = 20

out_ws.freeze_panes = 'A2'
out_ws.auto_filter.ref = f'A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(records)+1}'
out_wb.save(OUTPUT_FILE)

# ========== 7. 验证 ==========
wb_v = load_workbook(OUTPUT_FILE)
ws_v = wb_v.active
print(f'\n=== 完成 ===')
print(f'输出: {OUTPUT_FILE}')
print(f'月份: {MONTH} | 公司: {len(records)} | 总列数: {len(OUTPUT_COLUMNS)}')

# 显示公式列的前3行
formula_output_names = [f[1] for f in FORMULA_DEFS_INTERNAL]
print('\n前3行公式示例:')
for ri in range(2, min(5, len(records)+2)):
    company = ws_v.cell(row=ri, column=1).value
    parts = []
    for col_name in formula_output_names:
        if col_name in col_letter_of:
            ci = list(OUTPUT_COLUMNS).index(col_name) + 1
            cell = ws_v.cell(row=ri, column=ci)
            parts.append(f'{col_name}={cell.value}')
    print(f'  {company}:')
    for p in parts[:6]:  # 只显示前6个
        print(f'    {p}')
